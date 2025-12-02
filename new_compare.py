import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
import sys
import re

sys.stdout.reconfigure(encoding="utf-8")

# ---------------- CONFIG: EDIT THIS PER USE CASE ---------------- #

VALIDATION_CONFIG = {
    # 1. Column-level expectations
    "required_columns": [
        # "ID", "Date", "Amount", ...
    ],
    "expected_column_order": [
        # If you care about exact order, list them here in order
        # "ID", "Date", "Amount", ...
    ],
    # dtype strings can be: "int", "float", "str", "datetime", "bool"
    "expected_dtypes": {
        # "ID": "int",
        # "Amount": "float",
        # "Email": "str",
        # "CreatedAt": "datetime",
    },

    # 2. Row-level expectations
    "primary_key": None,  # e.g. "ID" if you have a unique key
    "expected_min_rows": None,  # e.g. 1_000
    "expected_max_rows": None,  # e.g. 2_000

    # 3. Missing data
    "max_missing_ratio_per_column": 0.2,  # 20% allowed, tweak as needed

    # 4. Value & constraint validation
    # ranges: {column: (min_allowed, max_allowed)} (None = no bound)
    "value_ranges": {
        # "Age": (0, 120),
        # "Amount": (0, None),
    },
    # regex patterns: {column: r"regex"}
    "regex_patterns": {
        # "Email": r"^[^@\s]+@[^@\s]+\.[^@\s]+$",
    },
    # allowed categories: {column: [list_of_allowed_values]}
    "allowed_categories": {
        # "Status": ["Active", "Inactive", "Pending"],
    },
    # uniqueness constraints: list of columns that must be unique
    "unique_columns": [
        # "ID",
    ],

    # 6. Business rules: each rule is a function df -> boolean Series
    "business_rules": {
        # "Total = Price * Quantity": lambda df: df["Total"] == df["Price"] * df["Quantity"],
    },

    # 7. File-level checks
    "filename_pattern": None,  # e.g. r"^Report_\d{8}\.xlsx$"
    "min_file_size_bytes": None,  # e.g. 1_000
    "max_file_size_bytes": None,
    "expected_sheet_names": None,  # e.g. ["Sheet1"]

    # 9. Cross-file checks / data drift can be added later
}

# ---------------- HELPER FUNCTIONS ---------------- #

def map_pd_dtype_to_expected(pd_dtype):
    """Rough mapping from pandas dtype to simplified type labels."""
    if pd.api.types.is_integer_dtype(pd_dtype):
        return "int"
    if pd.api.types.is_float_dtype(pd_dtype):
        return "float"
    if pd.api.types.is_bool_dtype(pd_dtype):
        return "bool"
    if pd.api.types.is_datetime64_any_dtype(pd_dtype):
        return "datetime"
    return "str"


def add_issue_row(ws, issue_type, row, col, details):
    """Append an issue row into ValidationIssues sheet."""
    ws.append([issue_type, row, col, details])


def validate_columns(df, summary, issues_ws, cfg):
    cols = list(df.columns)

    # 1.1 presence / unexpected
    required = cfg.get("required_columns") or []
    missing = [c for c in required if c not in cols]
    unexpected = [c for c in cols if required and c not in required]

    if missing:
        msg = f"✗ Missing required columns: {missing}"
        summary.append(msg)
        add_issue_row(issues_ws, "MissingColumn", "", "", msg)
    else:
        if required:
            summary.append("✓ All required columns present")

    if unexpected:
        msg = f"⚠ Unexpected columns present: {unexpected}"
        summary.append(msg)
        add_issue_row(issues_ws, "UnexpectedColumn", "", "", msg)

    # 1.2 order
    expected_order = cfg.get("expected_column_order") or []
    if expected_order:
        if cols == expected_order:
            summary.append("✓ Column order matches expected")
        else:
            summary.append("✗ Column order DOES NOT match expected")
            for idx, col in enumerate(cols):
                if col in expected_order:
                    expected_idx = expected_order.index(col)
                    if expected_idx != idx:
                        details = (
                            f"Column '{col}': expected index {expected_idx}, "
                            f"found at index {idx}"
                        )
                        add_issue_row(
                            issues_ws,
                            "ColumnOrderMismatch",
                            "",
                            col,
                            details,
                        )

    # 1.3 dtype validation
    expected_dtypes = cfg.get("expected_dtypes") or {}
    for col in expected_dtypes:
        if col not in df.columns:
            continue
        actual_pd_dtype = df[col].dtype
        actual_simple = map_pd_dtype_to_expected(actual_pd_dtype)
        expected_simple = expected_dtypes[col]
        if actual_simple != expected_simple:
            msg = (
                f"✗ Datatype mismatch for '{col}': expected {expected_simple}, "
                f"found {actual_simple} (pandas dtype {actual_pd_dtype})"
            )
            summary.append(msg)
            add_issue_row(issues_ws, "ColumnTypeMismatch", "", col, msg)


def validate_rows(df1, df2, summary, issues_ws, cfg):
    # 2.1 row count expectations vs config
    n_rows = len(df1)
    min_rows = cfg.get("expected_min_rows")
    max_rows = cfg.get("expected_max_rows")

    if min_rows is not None and n_rows < min_rows:
        msg = f"✗ Row count {n_rows} below expected minimum {min_rows}"
        summary.append(msg)
        add_issue_row(issues_ws, "RowCountBelowMin", "", "", msg)
    if max_rows is not None and n_rows > max_rows:
        msg = f"✗ Row count {n_rows} above expected maximum {max_rows}"
        summary.append(msg)
        add_issue_row(issues_ws, "RowCountAboveMax", "", "", msg)

    # 2.2 row order based on primary key
    pk = cfg.get("primary_key")
    if pk and pk in df1.columns and pk in df2.columns:
        if df1[pk].isnull().any() or df2[pk].isnull().any():
            summary.append(f"⚠ Primary key '{pk}' contains nulls; order check may be unreliable.")
        else:
            order1 = list(df1[pk])
            order2 = list(df2[pk])
            if order1 == order2:
                summary.append(f"✓ Row order MATCH based on primary key '{pk}'")
            else:
                summary.append(f"✗ Row order mismatch based on primary key '{pk}'")
                # Report a few mismatches
                for i, (v1, v2) in enumerate(zip(order1, order2)):
                    if v1 != v2:
                        details = (
                            f"Row {i+1}: SplashBI {pk}={v1}, Express {pk}={v2}"
                        )
                        add_issue_row(issues_ws, "RowOrderMismatch", i + 2, pk, details)

        # 2.3 missing row detection (present in df1 but not in df2)
        missing_in_df2 = set(df1[pk]) - set(df2[pk])
        missing_in_df1 = set(df2[pk]) - set(df1[pk])
        if missing_in_df2:
            msg = f"✗ {len(missing_in_df2)} rows present in SplashBI but missing in Express based on '{pk}'"
            summary.append(msg)
            for val in list(missing_in_df2)[:100]:  # cap to 100
                add_issue_row(
                    issues_ws,
                    "MissingRowInExpress",
                    "",
                    pk,
                    f"{pk}={val}",
                )
        if missing_in_df1:
            msg = f"✗ {len(missing_in_df1)} rows present in Express but missing in SplashBI based on '{pk}'"
            summary.append(msg)
            for val in list(missing_in_df1)[:100]:
                add_issue_row(
                    issues_ws,
                    "MissingRowInSplashBI",
                    "",
                    pk,
                    f"{pk}={val}",
                )


def validate_missing_data(df, summary, issues_ws, cfg):
    max_ratio = cfg.get("max_missing_ratio_per_column", 1.0)
    total_rows = len(df)
    if total_rows == 0:
        return

    summary.append("— Missing data per column —")
    for col in df.columns:
        missing_count = df[col].isna().sum()
        if missing_count == 0:
            continue
        ratio = missing_count / total_rows
        msg = f"Column '{col}': {missing_count} missing ({ratio:.2%})"
        summary.append(msg)
        add_issue_row(issues_ws, "MissingData", "", col, msg)
        if ratio > max_ratio:
            warn = (
                f"✗ Column '{col}' missing ratio {ratio:.2%} "
                f"exceeds threshold {max_ratio:.2%}"
            )
            summary.append(warn)
            add_issue_row(issues_ws, "MissingDataHigh", "", col, warn)


def validate_duplicates(df, summary, issues_ws, cfg):
    # 5.1 duplicate rows
    duplicate_rows = df.duplicated(keep=False)
    if duplicate_rows.any():
        count = duplicate_rows.sum()
        msg = f"✗ {count} duplicate rows found"
        summary.append(msg)
        for idx in df.index[duplicate_rows][:100]:  # cap
            add_issue_row(issues_ws, "DuplicateRow", idx + 2, "", "Duplicate row")
    else:
        summary.append("✓ No duplicate rows")

    # 5.2 duplicate primary keys / unique columns
    unique_cols = cfg.get("unique_columns") or []
    pk = cfg.get("primary_key")
    if pk and pk not in unique_cols:
        unique_cols.append(pk)

    for col in unique_cols:
        if col not in df.columns:
            continue
        dup = df[col].duplicated(keep=False)
        if dup.any():
            count = dup.sum()
            msg = f"✗ {count} duplicate values in '{col}' (must be unique)"
            summary.append(msg)
            for idx, val in zip(df.index[dup][:100], df.loc[dup, col].head(100)):
                add_issue_row(
                    issues_ws,
                    "DuplicateKey",
                    idx + 2,
                    col,
                    f"Value '{val}' duplicated",
                )
        else:
            summary.append(f"✓ Column '{col}' has unique values")


def validate_value_constraints(df, summary, issues_ws, cfg):
    # 4.1 ranges
    ranges = cfg.get("value_ranges") or {}
    for col, (min_val, max_val) in ranges.items():
        if col not in df.columns:
            continue
        series = df[col]
        if min_val is not None:
            mask = series < min_val
            if mask.any():
                count = mask.sum()
                msg = f"✗ {count} values in '{col}' below minimum {min_val}"
                summary.append(msg)
                for idx, val in zip(series[mask].index[:100], series[mask].head(100)):
                    add_issue_row(
                        issues_ws,
                        "ValueBelowMin",
                        idx + 2,
                        col,
                        f"{val} < {min_val}",
                    )
        if max_val is not None:
            mask = series > max_val
            if mask.any():
                count = mask.sum()
                msg = f"✗ {count} values in '{col}' above maximum {max_val}"
                summary.append(msg)
                for idx, val in zip(series[mask].index[:100], series[mask].head(100)):
                    add_issue_row(
                        issues_ws,
                        "ValueAboveMax",
                        idx + 2,
                        col,
                        f"{val} > {max_val}",
                    )

    # 4.2 regex
    regex_patterns = cfg.get("regex_patterns") or {}
    for col, pattern in regex_patterns.items():
        if col not in df.columns:
            continue
        regex = re.compile(pattern)
        mask = ~df[col].astype(str).fillna("").str.match(regex)
        mask &= df[col].notna()
        if mask.any():
            count = mask.sum()
            msg = f"✗ {count} values in '{col}' do not match regex '{pattern}'"
            summary.append(msg)
            for idx, val in zip(df.index[mask][:100], df.loc[mask, col].head(100)):
                add_issue_row(
                    issues_ws,
                    "RegexMismatch",
                    idx + 2,
                    col,
                    f"Value '{val}' fails pattern",
                )

    # 6. business rules
    business_rules = cfg.get("business_rules") or {}
    for rule_name, rule_fn in business_rules.items():
        try:
            mask_ok = rule_fn(df)
            mask_bad = ~mask_ok.fillna(False)
            if mask_bad.any():
                count = mask_bad.sum()
                msg = f"✗ {count} rows violate business rule: {rule_name}"
                summary.append(msg)
                for idx in df.index[mask_bad][:100]:
                    add_issue_row(
                        issues_ws,
                        "BusinessRuleViolation",
                        idx + 2,
                        "",
                        rule_name,
                    )
            else:
                summary.append(f"✓ Business rule OK: {rule_name}")
        except Exception as e:
            msg = f"⚠ Error evaluating business rule '{rule_name}': {e}"
            summary.append(msg)


def validate_file_level(path, summary, issues_ws, cfg):
    # 7.1 file name pattern
    filename = os.path.basename(path)
    pattern = cfg.get("filename_pattern")
    if pattern:
        if not re.match(pattern, filename):
            msg = f"✗ Filename '{filename}' does not match expected pattern '{pattern}'"
            summary.append(msg)
            add_issue_row(issues_ws, "FilenamePatternMismatch", "", "", msg)
        else:
            summary.append("✓ Filename pattern OK")

    # 7.2 file size
    size = os.path.getsize(path)
    min_size = cfg.get("min_file_size_bytes")
    max_size = cfg.get("max_file_size_bytes")
    if min_size is not None and size < min_size:
        msg = f"✗ File size {size} bytes below minimum {min_size}"
        summary.append(msg)
        add_issue_row(issues_ws, "FileTooSmall", "", "", msg)
    if max_size is not None and size > max_size:
        msg = f"✗ File size {size} bytes above maximum {max_size}"
        summary.append(msg)
        add_issue_row(issues_ws, "FileTooLarge", "", "", msg)

    # 7.3 sheet names
    expected_sheets = cfg.get("expected_sheet_names")
    if expected_sheets:
        try:
            xls = pd.ExcelFile(path)
            actual_sheets = xls.sheet_names
            missing = [s for s in expected_sheets if s not in actual_sheets]
            extra = [s for s in actual_sheets if s not in expected_sheets]
            if missing:
                msg = f"✗ Missing expected sheets: {missing}"
                summary.append(msg)
                add_issue_row(issues_ws, "MissingSheet", "", "", msg)
            if extra:
                msg = f"⚠ Extra sheets present: {extra}"
                summary.append(msg)
                add_issue_row(issues_ws, "ExtraSheet", "", "", msg)
            if not missing and not extra:
                summary.append("✓ Sheet names match expected")
        except Exception as e:
            summary.append(f"⚠ Could not inspect sheet names: {e}")


# ---------------- MAIN SCRIPT ---------------- #

# Hide Tkinter root window
Tk().withdraw()

print("Select SplashBI report:")
file1 = askopenfilename(title="Select SplashBI report", filetypes=[("Excel files","*.xlsx *.xls")])
if not file1:
    print("No file selected. Exiting...")
    exit()

print("Select Express report:")
file2 = askopenfilename(title="Select Express report", filetypes=[("Excel files","*.xlsx *.xls")])
if not file2:
    print("No file selected. Exiting...")
    exit()

df1 = pd.read_excel(file1)
df2 = pd.read_excel(file2)

summary = []

# 4. Column comparison between the two files (existing logic)
if list(df1.columns) == list(df2.columns):
    summary.append("✓ Columns MATCH between SplashBI and Express")
else:
    summary.append("✗ Columns DO NOT MATCH between SplashBI and Express")
    summary.append(f"SplashBI Columns: {list(df1.columns)}")
    summary.append(f"Express Columns: {list(df2.columns)}")

# 5. Row count comparison between files
if len(df1) == len(df2):
    summary.append("✓ Row Count MATCH between SplashBI and Express")
else:
    summary.append("✗ Row Count MISMATCH between SplashBI and Express")

# 6. Exact equality / row order
if df1.equals(df2):
    summary.append("✓ Row order + data MATCH exactly")
    row_order_mismatch = False
else:
    summary.append("✗ Row order OR data mismatch detected")
    row_order_mismatch = True

# 7. Numeric totals comparison
numeric_cols = df1.select_dtypes(include=["int64", "float64"]).columns
for col in numeric_cols:
    if col in df2.columns and df1[col].sum() == df2[col].sum():
        summary.append(f"✓ Total MATCH for: {col}")
    else:
        summary.append(f"✗ Total MISMATCH for: {col}")

# 8. Cell-by-cell comparison + output file
output_dir = os.path.join(os.path.dirname(file1), "output")
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, "mismatch_report.xlsx")

# Try to avoid permission error
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except PermissionError:
        print(
            f"✗ Cannot overwrite {output_file}.\n"
            "   Please close the file in Excel or any other program and run again."
        )
        exit()

df1.to_excel(output_file, index=False)
wb = load_workbook(output_file)
ws = wb.active
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

comparison = df1.eq(df2)
mismatch_found = row_order_mismatch

for r in range(len(df1)):
    for c in range(len(df1.columns)):
        if c >= df2.shape[1] or r >= df2.shape[0]:
            # Out-of-bounds in df2 also count as mismatch
            ws.cell(row=r + 2, column=c + 1).fill = red_fill
            mismatch_found = True
        else:
            if not comparison.iat[r, c]:
                ws.cell(row=r + 2, column=c + 1).fill = red_fill
                mismatch_found = True

# ------------- NEW: VALIDATION SHEET & CHECKS ------------- #

issues_ws = wb.create_sheet("ValidationIssues")
issues_ws.append(["IssueType", "RowNumber", "ColumnName", "Details"])

# Column-level validation (schema)
validate_columns(df1, summary, issues_ws, VALIDATION_CONFIG)

# Row-level validation between files (primary key, missing rows)
validate_rows(df1, df2, summary, issues_ws, VALIDATION_CONFIG)

# Missing data & duplicate checks (on SplashBI file; you can repeat for df2 if needed)
validate_missing_data(df1, summary, issues_ws, VALIDATION_CONFIG)
validate_duplicates(df1, summary, issues_ws, VALIDATION_CONFIG)

# Value/constraint/business rules
validate_value_constraints(df1, summary, issues_ws, VALIDATION_CONFIG)

# File-level validation (SplashBI file)
validate_file_level(file1, summary, issues_ws, VALIDATION_CONFIG)

wb.save(output_file)
wb.close()

# 9. Save summary text
summary_file = os.path.join(output_dir, "summary.txt")
with open(summary_file, "w", encoding="utf-8") as f:
    for line in summary:
        f.write(line + "\n")

# 10. Final message
if mismatch_found:
    print(f"✗ Differences Found → Check {output_file}")
else:
    print(f"✓ Both Files Match perfectly including row order")

print(f"Summary saved in {summary_file}")

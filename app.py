import streamlit as st
import pandas as pd
import json
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import base64

st.set_page_config(page_title="Excel Comparator & Validator", layout="wide")

# ---------------- DEFAULT CONFIG ---------------- #
DEFAULT_CONFIG = {
    "required_columns": [],
    "expected_column_order": [],
    "expected_dtypes": {},
    "primary_key": None,
    "expected_min_rows": None,
    "expected_max_rows": None,
    "max_missing_ratio_per_column": 0.2,
    "value_ranges": {},
    "regex_patterns": {},
    "allowed_categories": {},
    "unique_columns": [],
    "business_rules": {},
    "filename_pattern": None,
    "min_file_size_bytes": None,
    "max_file_size_bytes": None,
    "expected_sheet_names": None
}

# ---------------- HELPERS ---------------- #
def map_pd_dtype_to_expected(pd_dtype):
    if pd.api.types.is_integer_dtype(pd_dtype):
        return "int"
    if pd.api.types.is_float_dtype(pd_dtype):
        return "float"
    if pd.api.types.is_bool_dtype(pd_dtype):
        return "bool"
    if pd.api.types.is_datetime64_any_dtype(pd_dtype):
        return "datetime"
    return "str"


def add_issue_row_list(issue_list, issue_type, row, col, details):
    issue_list.append({"IssueType": issue_type, "RowNumber": row, "ColumnName": col, "Details": details})


# Validation helpers (kept but expanded for richer summaries)
def validate_columns(df, summary_items, issues_list, cfg):
    cols = list(df.columns)
    required = cfg.get("required_columns") or []
    missing = [c for c in required if c not in cols]
    unexpected = [c for c in cols if required and c not in required]

    summary_items.append({"section": "Columns", "message": f"Total columns in file: {len(cols)}"})
    if missing:
        summary_items.append({"section": "Columns", "message": f"Missing required columns ({len(missing)}): {missing}"})
        add_issue_row_list(issues_list, "MissingColumn", "", "", f"Missing: {missing}")
    else:
        if required:
            summary_items.append({"section": "Columns", "message": "All required columns present"})

    if unexpected:
        summary_items.append({"section": "Columns", "message": f"Unexpected columns ({len(unexpected)}): {unexpected}"})
        add_issue_row_list(issues_list, "UnexpectedColumn", "", "", f"Unexpected: {unexpected}")

    expected_order = cfg.get("expected_column_order") or []
    if expected_order:
        if cols == expected_order:
            summary_items.append({"section": "Columns", "message": "Column order matches expected"})
        else:
            # compute how many are out of place
            out_of_place = [c for i, c in enumerate(cols) if c in expected_order and expected_order.index(c) != i]
            summary_items.append({"section": "Columns", "message": f"Column order does not match expected. Out of place: {len(out_of_place)}"})
            for idx, col in enumerate(cols):
                if col in expected_order:
                    expected_idx = expected_order.index(col)
                    if expected_idx != idx:
                        details = f"Column '{col}': expected index {expected_idx}, found at index {idx}"
                        add_issue_row_list(issues_list, "ColumnOrderMismatch", "", col, details)

    expected_dtypes = cfg.get("expected_dtypes") or {}
    dtype_mismatches = []
    for col in expected_dtypes:
        if col not in df.columns:
            continue
        actual_pd_dtype = df[col].dtype
        actual_simple = map_pd_dtype_to_expected(actual_pd_dtype)
        expected_simple = expected_dtypes[col]
        if actual_simple != expected_simple:
            msg = (f"Datatype mismatch for '{col}': expected {expected_simple}, found {actual_simple} (pandas dtype {actual_pd_dtype})")
            dtype_mismatches.append(msg)
            add_issue_row_list(issues_list, "ColumnTypeMismatch", "", col, msg)
    if dtype_mismatches:
        summary_items.append({"section": "Columns", "message": f"Datatype mismatches: {len(dtype_mismatches)}"})
        for m in dtype_mismatches:
            summary_items.append({"section": "Columns", "message": m})
    else:
        summary_items.append({"section": "Columns", "message": "Column datatypes match expected (where provided)"})


def validate_rows(df1, df2, summary_items, issues_list, cfg):
    n_rows = len(df1)
    min_rows = cfg.get("expected_min_rows")
    max_rows = cfg.get("expected_max_rows")

    summary_items.append({"section": "Rows", "message": f"Row count (SplashBI): {n_rows}, (Express): {len(df2)}"})

    if min_rows is not None and n_rows < min_rows:
        msg = f"Row count {n_rows} below expected minimum {min_rows}"
        summary_items.append({"section": "Rows", "message": msg})
        add_issue_row_list(issues_list, "RowCountBelowMin", "", "", msg)
    if max_rows is not None and n_rows > max_rows:
        msg = f"Row count {n_rows} above expected maximum {max_rows}"
        summary_items.append({"section": "Rows", "message": msg})
        add_issue_row_list(issues_list, "RowCountAboveMax", "", "", msg)

    pk = cfg.get("primary_key")
    if pk and pk in df1.columns and pk in df2.columns:
        nulls1 = df1[pk].isnull().sum()
        nulls2 = df2[pk].isnull().sum()
        if nulls1 or nulls2:
            summary_items.append({"section": "Rows", "message": f"Primary key '{pk}' nulls: SplashBI={nulls1}, Express={nulls2}"})
            add_issue_row_list(issues_list, "PrimaryKeyNulls", "", pk, f"SplashBI nulls={nulls1}, Express nulls={nulls2}")
        else:
            order1 = list(df1[pk])
            order2 = list(df2[pk])
            if order1 == order2:
                summary_items.append({"section": "Rows", "message": f"Row order MATCH based on primary key '{pk}'"})
            else:
                # compute how many rows differ in the same positions
                diffs = sum(1 for a, b in zip(order1, order2) if a != b)
                summary_items.append({"section": "Rows", "message": f"Row order mismatch based on '{pk}': {diffs} rows in different positions"})
                for i, (v1, v2) in enumerate(zip(order1, order2)):
                    if v1 != v2 and len(issues_list) < 500:
                        details = f"Row {i+1}: SplashBI {pk}={v1}, Express {pk}={v2}"
                        add_issue_row_list(issues_list, "RowOrderMismatch", i + 2, pk, details)

        missing_in_df2 = set(df1[pk]) - set(df2[pk])
        missing_in_df1 = set(df2[pk]) - set(df1[pk])
        if missing_in_df2:
            summary_items.append({"section": "Rows", "message": f"Rows present in SplashBI but missing in Express: {len(missing_in_df2)}"})
            for val in list(missing_in_df2)[:100]:
                add_issue_row_list(issues_list, "MissingRowInExpress", "", pk, f"{pk}={val}")
        if missing_in_df1:
            summary_items.append({"section": "Rows", "message": f"Rows present in Express but missing in SplashBI: {len(missing_in_df1)}"})
            for val in list(missing_in_df1)[:100]:
                add_issue_row_list(issues_list, "MissingRowInSplashBI", "", pk, f"{pk}={val}")


def validate_missing_data(df, summary_items, issues_list, cfg):
    max_ratio = cfg.get("max_missing_ratio_per_column", 1.0)
    total_rows = len(df)
    if total_rows == 0:
        return
    total_missing_cells = int(df.isna().sum().sum())
    summary_items.append({"section": "MissingData", "message": f"Total missing cells: {total_missing_cells} ({total_rows * len(df.columns)} possible)"})

    cols_exceeding = []
    for col in df.columns:
        missing_count = int(df[col].isna().sum())
        if missing_count == 0:
            continue
        ratio = missing_count / total_rows
        summary_items.append({"section": "MissingData", "message": f"{col}: {missing_count} missing ({ratio:.2%})"})
        add_issue_row_list(issues_list, "MissingData", "", col, f"{missing_count} missing ({ratio:.2%})")
        if ratio > max_ratio:
            cols_exceeding.append((col, missing_count, ratio))
            warn = (f"Column '{col}' missing ratio {ratio:.2%} exceeds threshold {max_ratio:.2%}")
            summary_items.append({"section": "MissingData", "message": warn})
            add_issue_row_list(issues_list, "MissingDataHigh", "", col, warn)
    if cols_exceeding:
        summary_items.append({"section": "MissingData", "message": f"Columns exceeding missing threshold: {len(cols_exceeding)}"})


def validate_duplicates(df, summary_items, issues_list, cfg):
    duplicate_rows = df.duplicated(keep=False)
    if duplicate_rows.any():
        count = int(duplicate_rows.sum())
        summary_items.append({"section": "Duplicates", "message": f"Duplicate rows found: {count}"})
        for idx in df.index[duplicate_rows][:100]:
            add_issue_row_list(issues_list, "DuplicateRow", int(idx) + 2, "", "Duplicate row")
    else:
        summary_items.append({"section": "Duplicates", "message": "No duplicate rows"})

    unique_cols = list(cfg.get("unique_columns") or [])
    pk = cfg.get("primary_key")
    if pk and pk not in unique_cols:
        unique_cols.append(pk)

    for col in unique_cols:
        if col not in df.columns:
            continue
        dup = df[col].duplicated(keep=False)
        if dup.any():
            count = int(dup.sum())
            summary_items.append({"section": "Duplicates", "message": f"{count} duplicate values in '{col}'"})
            for idx, val in zip(df.index[dup][:100], df.loc[dup, col].head(100)):
                add_issue_row_list(issues_list, "DuplicateKey", int(idx) + 2, col, f"Value '{val}' duplicated")
        else:
            summary_items.append({"section": "Duplicates", "message": f"Column '{col}' has unique values"})


def validate_value_constraints(df, summary_items, issues_list, cfg):
    ranges = cfg.get("value_ranges") or {}
    for col, bounds in ranges.items():
        if col not in df.columns:
            continue
        min_val, max_val = bounds if isinstance(bounds, (list, tuple)) and len(bounds) == 2 else (None, None)
        series = df[col]
        if min_val is not None:
            mask = series < min_val
            if mask.any():
                count = int(mask.sum())
                summary_items.append({"section": "ValueRanges", "message": f"{count} values in '{col}' below minimum {min_val}"})
                for idx, val in zip(series[mask].index[:100], series[mask].head(100)):
                    add_issue_row_list(issues_list, "ValueBelowMin", int(idx) + 2, col, f"{val} < {min_val}")
        if max_val is not None:
            mask = series > max_val
            if mask.any():
                count = int(mask.sum())
                summary_items.append({"section": "ValueRanges", "message": f"{count} values in '{col}' above maximum {max_val}"})
                for idx, val in zip(series[mask].index[:100], series[mask].head(100)):
                    add_issue_row_list(issues_list, "ValueAboveMax", int(idx) + 2, col, f"{val} > {max_val}")

    regex_patterns = cfg.get("regex_patterns") or {}
    for col, pattern in regex_patterns.items():
        if col not in df.columns:
            continue
        try:
            regex = re.compile(pattern)
        except Exception as e:
            summary_items.append({"section": "Regex", "message": f"Invalid regex for '{col}': {e}"})
            continue
        mask = ~df[col].astype(str).fillna("").str.match(regex)
        mask &= df[col].notna()
        if mask.any():
            count = int(mask.sum())
            summary_items.append({"section": "Regex", "message": f"{count} values in '{col}' do not match regex"})
            for idx, val in zip(df.index[mask][:100], df.loc[mask, col].head(100)):
                add_issue_row_list(issues_list, "RegexMismatch", int(idx) + 2, col, f"Value '{val}' fails pattern")

    business_rules = cfg.get("business_rules") or {}
    for rule_name, rule_def in business_rules.items():
        summary_items.append({"section": "BusinessRules", "message": f"Business rule '{rule_name}' present but cannot be evaluated in UI."})


def validate_file_level(uploaded_file, filename, summary_items, issues_list, cfg):
    pattern = cfg.get("filename_pattern")
    if pattern:
        if not re.match(pattern, filename):
            msg = f"Filename '{filename}' does not match expected pattern '{pattern}'"
            summary_items.append({"section": "File", "message": msg})
            add_issue_row_list(issues_list, "FilenamePatternMismatch", "", "", msg)
        else:
            summary_items.append({"section": "File", "message": "Filename pattern OK"})

    size = len(uploaded_file.getvalue()) if hasattr(uploaded_file, "getvalue") else None
    min_size = cfg.get("min_file_size_bytes")
    max_size = cfg.get("max_file_size_bytes")
    if min_size is not None and size is not None and size < min_size:
        msg = f"File size {size} bytes below minimum {min_size}"
        summary_items.append({"section": "File", "message": msg})
        add_issue_row_list(issues_list, "FileTooSmall", "", "", msg)
    if max_size is not None and size is not None and size > max_size:
        msg = f"File size {size} bytes above maximum {max_size}"
        summary_items.append({"section": "File", "message": msg})
        add_issue_row_list(issues_list, "FileTooLarge", "", "", msg)

    expected_sheets = cfg.get("expected_sheet_names")
    if expected_sheets:
        try:
            xls = pd.ExcelFile(uploaded_file)
            actual_sheets = xls.sheet_names
            missing = [s for s in expected_sheets if s not in actual_sheets]
            extra = [s for s in actual_sheets if s not in expected_sheets]
            if missing:
                msg = f"Missing expected sheets: {missing}"
                summary_items.append({"section": "File", "message": msg})
                add_issue_row_list(issues_list, "MissingSheet", "", "", msg)
            if extra:
                msg = f"Extra sheets present: {extra}"
                summary_items.append({"section": "File", "message": msg})
                add_issue_row_list(issues_list, "ExtraSheet", "", "", msg)
            if not missing and not extra:
                summary_items.append({"section": "File", "message": "Sheet names match expected"})
        except Exception as e:
            summary_items.append({"section": "File", "message": f"Could not inspect sheet names: {e}"})


# ---------------- UI ---------------- #
st.title("Excel Comparator & Validator (SplashBI vs Express) — Enhanced Summary")
st.markdown("Upload two Excel reports and get a rich validation summary and downloadable mismatch workbook.")

with st.sidebar:
    st.header("Validation Config (JSON)")
    cfg_text = st.text_area("Edit JSON config (leave empty to use defaults)", height=320,
                            value=json.dumps(DEFAULT_CONFIG, indent=2))
    try:
        VALIDATION_CONFIG = json.loads(cfg_text) if cfg_text.strip() else DEFAULT_CONFIG
    except Exception as e:
        st.error(f"Invalid JSON in config: {e}")
        st.stop()

st.write("### 1) Upload files")
col1, col2 = st.columns(2)
with col1:
    uploaded1 = st.file_uploader("Select SplashBI report (xlsx/xls)", type=["xlsx", "xls"], key="splash")
with col2:
    uploaded2 = st.file_uploader("Select Express report (xlsx/xls)", type=["xlsx", "xls"], key="express")

# Prepare placeholders for sheet selections
xl1_selected_sheet = None
xl2_selected_sheet = None

# Only try to inspect sheets if file(s) uploaded
if uploaded1:
    try:
        bytes1 = uploaded1.read()
        xls1 = pd.ExcelFile(BytesIO(bytes1))
        xl1_sheets = xls1.sheet_names
        xl1_selected_sheet = st.selectbox("Select a sheet for Excel File 1 (SplashBI)", xl1_sheets, key="sheet1")
    except Exception as e:
        st.error(f"Could not read SplashBI file sheets: {e}")

if uploaded2:
    try:
        bytes2 = uploaded2.read()
        xls2 = pd.ExcelFile(BytesIO(bytes2))
        xl2_sheets = xls2.sheet_names
        xl2_selected_sheet = st.selectbox("Select a sheet for Excel File 2 (Express)", xl2_sheets, key="sheet2")
    except Exception as e:
        st.error(f"Could not read Express file sheets: {e}")

run_btn = st.button("Run Validation & Compare")

if run_btn:
    if not uploaded1 or not uploaded2:
        st.warning("Please upload both files.")
        st.stop()

    # ensure we still have sheet selection; fallback to first sheet if user didn't select
    try:
        # if we earlier read the uploads for sheet lists, bytes1/bytes2 exist; if not, read now
        if 'bytes1' not in locals():
            bytes1 = uploaded1.read()
        if 'bytes2' not in locals():
            bytes2 = uploaded2.read()

        with st.spinner("Reading Excel files..."):
            # create ExcelFile from fresh BytesIO objects so stream positions are independent
            xl1 = pd.ExcelFile(BytesIO(bytes1))
            xl2 = pd.ExcelFile(BytesIO(bytes2))
            # fallback to first sheet if user didn't choose
            if not xl1_selected_sheet:
                xl1_selected_sheet = xl1.sheet_names[0]
            if not xl2_selected_sheet:
                xl2_selected_sheet = xl2.sheet_names[0]

            # read sheets using fresh BytesIO objects
            df1 = pd.read_excel(BytesIO(bytes1), sheet_name=xl1_selected_sheet)
            df2 = pd.read_excel(BytesIO(bytes2), sheet_name=xl2_selected_sheet)
    except Exception as e:
        st.error(f"Error reading files: {e}")
        st.stop()

    summary_items = []
    issues_list = []

    # Basic column/row comparison notes
    if list(df1.columns) == list(df2.columns):
        summary_items.append({"section": "Compare", "message": "Columns MATCH between SplashBI and Express"})
    else:
        summary_items.append({"section": "Compare", "message": "Columns DO NOT MATCH between SplashBI and Express"})
        summary_items.append({"section": "Compare", "message": f"SplashBI Columns: {list(df1.columns)}"})
        summary_items.append({"section": "Compare", "message": f"Express Columns: {list(df2.columns)}"})

    if len(df1) == len(df2):
        summary_items.append({"section": "Compare", "message": "Row Count MATCH between SplashBI and Express"})
    else:
        summary_items.append({"section": "Compare", "message": f"Row Count MISMATCH: SplashBI={len(df1)}, Express={len(df2)}"})

    # Exact equality + detailed cell diffs
    equal = df1.equals(df2)
    if equal:
        summary_items.append({"section": "Compare", "message": "Row order + data MATCH exactly"})
    else:
        summary_items.append({"section": "Compare", "message": "Row order OR data mismatch detected"})

    # Numeric totals with diffs
    numeric_cols = df1.select_dtypes(include=["int64", "float64"]).columns
    numeric_summary = []
    for col in numeric_cols:
        if col in df2.columns:
            s1 = df1[col].sum(skipna=True)
            s2 = df2[col].sum(skipna=True)
            if pd.isna(s1) or pd.isna(s2):
                numeric_summary.append((col, s1, s2, None))
            else:
                diff = s1 - s2
                pct = (diff / s2 * 100) if s2 != 0 else None
                numeric_summary.append((col, s1, s2, diff, pct))
                if diff == 0:
                    summary_items.append({"section": "NumericTotals", "message": f"Total MATCH for: {col} = {s1}"})
                else:
                    summary_items.append({"section": "NumericTotals", "message": f"Total MISMATCH for: {col} — SplashBI={s1}, Express={s2}, Diff={diff}, PctDiff={pct if pct is not None else 'N/A'}"})

    # Prepare mismatch workbook with colored cells and detailed mismatch listing
    out_bytes = BytesIO()
    df1.to_excel(out_bytes, index=False, sheet_name="SplashBI")
    out_bytes.seek(0)
    wb = load_workbook(out_bytes)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Align shapes for comparison
    max_rows = max(df1.shape[0], df2.shape[0])
    max_cols = max(df1.shape[1], df2.shape[1])

    # Ensure df2 has same columns order as df1 for cell-by-cell compare where possible
    df2_aligned = df2.reindex(columns=df1.columns)

    # collect detailed cell diffs (limit to avoid huge memory)
    cell_mismatches = []
    for r in range(max_rows):
        for c in range(max_cols):
            try:
                v1 = df1.iat[r, c] if r < df1.shape[0] and c < df1.shape[1] else None
            except Exception:
                v1 = None
            try:
                v2 = df2_aligned.iat[r, c] if r < df2_aligned.shape[0] and c < df2_aligned.shape[1] else None
            except Exception:
                v2 = None
            if pd.isna(v1) and pd.isna(v2):
                continue
            if (v1 != v2) and not (pd.isna(v1) and pd.isna(v2)):
                # mark cell in workbook if within bounds
                if r < ws.max_row and c < ws.max_column:
                    try:
                        ws.cell(row=r + 2, column=c + 1).fill = red_fill
                    except Exception:
                        pass
                cell_mismatches.append({
                    "row": r + 2,
                    "col_index": c + 1,
                    "column": df1.columns[c] if c < len(df1.columns) else f"col_{c}",
                    "splash_value": v1,
                    "express_value": v2
                })
                if len(cell_mismatches) >= 1000:
                    break
        if len(cell_mismatches) >= 1000:
            break

    mismatch_found = (not equal) or len(cell_mismatches) > 0

    # Append ValidationIssues sheet with structured issues and also add cell-level mismatches
    issues_ws = wb.create_sheet("ValidationIssues")
    issues_ws.append(["IssueType", "RowNumber", "ColumnName", "Details"]) 

    # Run in-depth validations for SplashBI (df1)
    validate_columns(df1, summary_items, issues_list, VALIDATION_CONFIG)
    validate_rows(df1, df2, summary_items, issues_list, VALIDATION_CONFIG)
    validate_missing_data(df1, summary_items, issues_list, VALIDATION_CONFIG)
    validate_duplicates(df1, summary_items, issues_list, VALIDATION_CONFIG)
    validate_value_constraints(df1, summary_items, issues_list, VALIDATION_CONFIG)
    validate_file_level(uploaded1, getattr(uploaded1, "name", "uploaded1.xlsx"), summary_items, issues_list, VALIDATION_CONFIG)

    # write issues_list
    for issue in issues_list:
        issues_ws.append([issue["IssueType"], issue["RowNumber"], issue["ColumnName"], issue["Details"]])

    # also append cell-level mismatches (as separate rows)
    if cell_mismatches:
        issues_ws.append(["CellMismatch", "RowNumber", "ColumnName", "SplashValue || ExpressValue"])
        for cm in cell_mismatches:
            issues_ws.append(["CellMismatch", cm["row"], cm["column"], f"{cm['splash_value']} || {cm['express_value']}"])

    # Summary sheet (human readable) for quick glance
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Section", "Message"])
    for item in summary_items:
        summary_ws.append([item.get("section"), item.get("message")])

    # Save workbook
    out_stream = BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)

    # Also prepare a plain text summary with counts and top mismatches
    top_cells = cell_mismatches[:20]
    text_lines = []
    text_lines.append("=== Quick Validation Summary ===")
    for it in summary_items:
        text_lines.append(f"[{it.get('section')}] {it.get('message')}")

    text_lines.append("\nTop cell mismatches (first 20):")
    for cm in top_cells:
        text_lines.append(f"Row {cm['row']}, Col '{cm['column']}': SplashBI={cm['splash_value']} | Express={cm['express_value']}")

    text_lines.append(f"\nTotal cell mismatches found (capped at 1000): {len(cell_mismatches)}")
    text_lines.append(f"Total validation issues recorded: {len(issues_list)}")

    summary_text = "\n".join(text_lines)
    summary_bytes = summary_text.encode("utf-8")

    # Show results
    st.subheader("Summary")
    if mismatch_found:
        st.error("Differences Found — see downloads and " + "'ValidationIssues' sheet within the mismatch workbook")
    else:
        st.success("Both Files Match perfectly including row order")

    st.code(summary_text)

    st.subheader("Validation Issues (table)")
    if issues_list or cell_mismatches:
        # combine for display
        display_rows = issues_list.copy()
        # convert cell mismatches to same shape for the dataframe display
        for cm in cell_mismatches[:500]:
            display_rows.append({"IssueType": "CellMismatch", "RowNumber": cm["row"], "ColumnName": cm["column"], "Details": f"Splash: {cm['splash_value']} || Express: {cm['express_value']}"})
        df_issues = pd.DataFrame(display_rows)
        st.dataframe(df_issues)
    else:
        st.info("No validation issues recorded.")

    # Provide downloads
    st.subheader("Downloads")
    st.download_button("Download mismatch_report.xlsx", data=out_stream.getvalue(), file_name="mismatch_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.download_button("Download summary.txt", data=summary_bytes, file_name="summary.txt", mime="text/plain")

    # Previews
    with st.expander("Preview: SplashBI (first 5 rows)"):
        st.dataframe(df1.head())
    with st.expander("Preview: Express (first 5 rows)"):
        st.dataframe(df2.head())

    st.balloons()
